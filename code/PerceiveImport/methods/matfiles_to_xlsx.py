""" insert selected .mat files to an existing PerceiveMetadata.xlsx sheet """

import openpyxl
from openpyxl import Workbook, load_workbook

import PerceiveImport.methods.select_matfiles as matfiles



def insert_matfiles_to_Excel(sub, rec_modality):
    """ insert_matfiles_to_Excel() method:
    This method inserts all matfiles from a chosen subject (e.g. "sub-021) to the first column of the Perceive_Metadata Excel sheet.
    Choose the recording modality ("rec_modality" = "Streaming", "Survey", "Timeline") of your choice.

    The matfilenames will be inserted after the last edited row of the .xlsx sheet.
    The Perceive_Metadata.xlsx file be saved.
    
    """

    # load existing .xlsx file
    wb = load_workbook('Perceive_Metadata.xlsx')
    ws = wb.active # this gets the current active worksheet

    # list I want to append to a specific column
    matfile_list, _ = matfiles.select_matfiles(sub, rec_modality) 
    
    # find the max row number from your Excel sheet
    max_rows = ws.max_row

    # define the start in +1 row after the last row in PerceiveMetadata.xlsx
    row_start = max_rows + 1
    column_index = 1 # define the column where to insert list

    for i, value in enumerate(matfile_list, start=row_start):
        ws.cell(row=i, column=column_index).value = value

    wb.save("Perceive_Metadata.xlsx")


# def insert_sub_to_Excel():
#     wb = load_workbook('Perceive_Metadata.xlsx')
#     ws = wb.active # this gets the current active worksheet



#     wb.save("Perceive_Metadata.xlsx")



# def insert_recmod_to_Excel():
#     wb = load_workbook('Perceive_Metadata.xlsx')
#     ws = wb.active # this gets the current active worksheet

#     rec_modality_dict = {
#         "Survey": "LMTD",
#         "Streaming": "BrainSense",
#         "Timeline": "CHRONIC"
#         }
    
#     for file in ws.iter_cols(min_col=1, max_col=1): # loop through every filename in column 1
#         if rec_modality_dict.values() in file:
#             print(rec_modality_dict.keys) # how can I specify to only get a specific key to a value in file???
    
#     # how can I add a specific key from rec_modality_dict to column 3 in a specific row??

#     wb.save("Perceive_Metadata.xlsx")
